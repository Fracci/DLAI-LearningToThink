"""
plot_transfer.py — headline figures for the transfer (fine-tuning) results.

Reads full_results.xlsx (raw per-seed/epoch EM & PD), results_summary.xlsx
(paired gap-over-baseline), and the per-seed training logs (loss). Produces the
training-evolution trajectories — OOD accuracy vs epoch per arm against the shared
random baseline with seed bands — plus the loss curves and the 6-digit gap
spectrum. Headline metric is exact-match (EM); per-digit (PD) is the diagnostic.
Noisy curves are drawn faint with a bold smoothed overlay for readability.
"""
import os
import glob
import csv
import statistics as st
from collections import defaultdict
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
from openpyxl import load_workbook

OUTDIR = "Plots/TrainingResults"
INDIR = "Results/Training"
INDIR_LOSS = "Results/Logs"
DPI = 300
FULL = "full_results.xlsx"
SUMM = "results_summary.xlsx"
SMOOTH_W = 4          # centered moving-average window for the bold overlay (odd)

COLORS = {
    "Rule30 (A)":   "#d1495b",
    "Rollout (A)":  "#6a3d9a",
    "Carry (A)":    "#1b9e77",
    "Baseline (B)": "#8a8a8a",
}
ARM_ORDER = ["Rule30 (A)", "Rollout (A)", "Carry (A)", "Baseline (B)"]
PRETTY = lambda m: m.replace(" (A)", "").replace(" (B)", " (baseline)")

# Per-seed training logs for the LOSS figure (loss isn't in the xlsx). Edit the
# glob patterns / loss column to match your saved logs; Baseline loss comes from
# the A/B run's loss_B column.
LOSS_LOGS = {
    "Rule30 (A)":   ("Rule30_seed*_log.csv",    "loss_A"),
    "Rollout (A)":  ("rollout_seed*_log.csv",   "loss_A"),
    "Carry (A)":    ("carryonly_seed*_log.csv", "loss_A"),
    "Baseline (B)": ("seed*_log.csv",           "loss_B"),
}

plt.rcParams.update({
    "font.family": "DejaVu Sans", "font.size": 14,
    "axes.titlesize": 16, "axes.titleweight": "bold", "axes.labelsize": 14,
    "xtick.labelsize": 12, "ytick.labelsize": 12,
    "axes.spines.top": False, "axes.spines.right": False,
    "axes.grid": True, "grid.alpha": 0.22, "grid.linewidth": 0.8,
    "legend.frameon": False, "legend.fontsize": 12.5, "figure.dpi": 120,
})


def smooth(ys, w=SMOOTH_W):
    """Centered moving average; preserves length, shrinks window at the edges."""
    if w <= 1 or len(ys) < 3:
        return ys
    half = w // 2
    out = []
    for i in range(len(ys)):
        lo = max(0, i - half); hi = min(len(ys), i + half + 1)
        out.append(sum(ys[lo:hi]) / (hi - lo))
    return out


def load_all_results():
    """Read All_Results -> [model][eval_set][metric] -> {epoch: [vals over seeds]}."""
    wb = load_workbook(os.path.join(INDIR, FULL), read_only=True)
    ws = wb["All_Results"]
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list))))
    for model, seed, epoch, eset, em, pd in ws.iter_rows(min_row=2, values_only=True):
        if model is None:
            continue
        data[model][eset]["EM"][int(epoch)].append(float(em))
        data[model][eset]["PD"][int(epoch)].append(float(pd))
    return data


def mean_std_by_epoch(epoch_map):
    eps = sorted(epoch_map)
    means = [st.mean(epoch_map[e]) for e in eps]
    stds = [st.pstdev(epoch_map[e]) if len(epoch_map[e]) > 1 else 0.0 for e in eps]
    return eps, means, stds


def _plot_arm(ax, eps, means, stds, color, baseline=False, band=True):
    """Faint raw mean + bold smoothed line + optional seed band, for readability."""
    ls = "--" if baseline else "-"
    sm = smooth(means)
    if band:
        lo = [m - s for m, s in zip(means, stds)]; hi = [m + s for m, s in zip(means, stds)]
        ax.fill_between(eps, lo, hi, color=color, alpha=0.10, linewidth=0)
    ax.plot(eps, means, ls=ls, color=color, lw=1.0, alpha=0.30)          # faint raw
    ax.plot(eps, sm, ls=ls, color=color, lw=3.0, label=PRETTY_LABEL(color))  # bold smoothed


def PRETTY_LABEL(color):
    for m, c in COLORS.items():
        if c == color:
            return PRETTY(m)
    return ""


def fig_trajectories(data, metric="EM"):
    """2x2: accuracy vs epoch for id/5/6/7-digit, all arms, faint-raw + bold-smoothed."""
    eval_sets = ["id", "5dig", "6dig", "7dig"]
    titles = {"id": "In-distribution (3–4 digit)", "5dig": "5-digit (OOD)",
              "6dig": "6-digit (OOD)", "7dig": "7-digit (OOD)"}
    fig, axes = plt.subplots(2, 2, figsize=(15, 10))
    for ax, es in zip(axes.flat, eval_sets):
        for model in ARM_ORDER:
            if es not in data[model]:
                continue
            eps, means, stds = mean_std_by_epoch(data[model][es][metric])
            _plot_arm(ax, eps, means, stds, COLORS[model], baseline=(model == "Baseline (B)"))
        ax.set_title(titles[es]); ax.set_xlabel("epoch"); ax.set_ylabel(f"{metric} (%)")
        ax.set_ylim(-2, 103); ax.xaxis.set_major_locator(MaxNLocator(6))
        if es == "id":
            ax.legend(loc="lower right")
    fig.suptitle(f"OOD length generalization during fine-tuning — {metric}  "
                 f"(bold = smoothed, band = ±seed std)", fontsize=18, fontweight="bold", y=0.995)
    fig.tight_layout()
    fig.savefig(os.path.join(OUTDIR, f"tr_fig_trajectories_{metric}.png"), dpi=DPI, bbox_inches="tight")
    plt.close(fig)


def fig_6dig_focus(data):
    """The decisive length: 6-digit EM and PD vs epoch, side by side."""
    fig, (axE, axP) = plt.subplots(1, 2, figsize=(15, 5.8))
    for ax, metric in ((axE, "EM"), (axP, "PD")):
        for model in ARM_ORDER:
            eps, means, stds = mean_std_by_epoch(data[model]["6dig"][metric])
            _plot_arm(ax, eps, means, stds, COLORS[model], baseline=(model == "Baseline (B)"))
        ax.set_title(f"6-digit {metric}"); ax.set_xlabel("epoch"); ax.set_ylabel(f"{metric} (%)")
        ax.xaxis.set_major_locator(MaxNLocator(6))
    axE.legend(loc="upper left")
    fig.suptitle("The decisive length: 6-digit OOD generalization  (bold = smoothed, band = ±seed std)",
                 fontsize=17, fontweight="bold", y=1.0)
    fig.tight_layout()
    fig.savefig(os.path.join(OUTDIR, "tr_fig_6dig_focus.png"), dpi=DPI, bbox_inches="tight")
    plt.close(fig)


def fig_indist_mastery(data):
    """In-distribution EM convergence — the control that the OOD gap is pure length-gen."""
    fig, ax = plt.subplots(figsize=(10, 6))
    for model in ARM_ORDER:
        eps, means, stds = mean_std_by_epoch(data[model]["id"]["EM"])
        ax.plot(eps, smooth(means), ("--" if model == "Baseline (B)" else "-"),
                color=COLORS[model], lw=3.0, label=PRETTY(model))
    ax.set_title("In-distribution mastery during fine-tuning\nall arms saturate → the OOD gap is pure length generalization")
    ax.set_xlabel("epoch"); ax.set_ylabel("in-dist EM (%)")
    ax.set_ylim(-2, 103); ax.xaxis.set_major_locator(MaxNLocator(8))
    ax.legend(loc="lower right")
    fig.tight_layout()
    fig.savefig(os.path.join(OUTDIR, "tr_fig_indist_mastery.png"), dpi=DPI, bbox_inches="tight")
    plt.close(fig)


def load_loss():
    """Read per-seed training logs -> {model: {epoch: [loss over seeds]}}. Empty if no logs."""
    out = {}
    for model, (pat, col) in LOSS_LOGS.items():
        files = sorted(glob.glob(os.path.join(INDIR, pat)))
        if not files:
            continue
        epoch_map = defaultdict(list)
        for fn in files:
            try:
                rdr = csv.DictReader(open(fn))
                if col not in (rdr.fieldnames or []):
                    continue
                for r in rdr:
                    try:
                        epoch_map[int(r["epoch"])].append(float(r[col]))
                    except (ValueError, KeyError):
                        continue
            except OSError:
                continue
        if epoch_map:
            out[model] = epoch_map
    return out


def fig_loss(loss):
    """Training loss vs epoch per arm (log-scale y), mean ± seed std."""
    if not loss:
        print("  [loss] no seed logs found — skipping loss figure. "
              "Set LOSS_LOGS globs to your seed{N}_log.csv files.")
        return
    fig, ax = plt.subplots(figsize=(10, 6))
    for model in ARM_ORDER:
        if model not in loss:
            continue
        eps, means, stds = mean_std_by_epoch(loss[model])
        c = COLORS[model]; ls = "--" if model == "Baseline (B)" else "-"
        ax.plot(eps, means, ls=ls, color=c, lw=3.0, label=PRETTY(model))
        lo = [max(m - s, 1e-6) for m, s in zip(means, stds)]
        hi = [m + s for m, s in zip(means, stds)]
        ax.fill_between(eps, lo, hi, color=c, alpha=0.12, linewidth=0)
    ax.set_yscale("log")
    ax.set_title("Training loss during fine-tuning\n(mean ± seed std, log scale)")
    ax.set_xlabel("epoch"); ax.set_ylabel("cross-entropy loss")
    ax.xaxis.set_major_locator(MaxNLocator(8))
    ax.legend(loc="upper right")
    fig.tight_layout()
    fig.savefig(os.path.join(OUTDIR, "tr_fig_loss.png"), dpi=DPI, bbox_inches="tight")
    plt.close(fig)


def load_gaps():
    wb = load_workbook(os.path.join(INDIR, SUMM), read_only=True)
    ws = wb["Gap_vs_Baseline"]; out = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        model, ood, metric, A, B, gmean, gstd, maxgap = row[:8]
        if model is None or model == "Model":
            continue
        out[(model, ood, metric)] = (float(str(gmean).replace("+", "")), float(gstd))
    return out


def fig_gap_spectrum(gaps):
    """Headline bars: 6-digit EM and PD gap over baseline as the compatibility spectrum."""
    arms = ["Rule30 (A)", "Rollout (A)", "Carry (A)"]
    pretty = {"Rule30 (A)": "Rule30\n(local)", "Rollout (A)": "Rollout\n(mismatched)",
              "Carry (A)": "Carry\n(matched)"}
    fig, (axE, axP) = plt.subplots(1, 2, figsize=(13.5, 5.8))
    for ax, metric in ((axE, "EM"), (axP, "PD")):
        vals = [gaps[(a, "6dig", metric)][0] for a in arms]
        errs = [gaps[(a, "6dig", metric)][1] for a in arms]
        colors = [COLORS[a] for a in arms]
        bars = ax.bar(range(len(arms)), vals, yerr=errs, capsize=7,
                      color=colors, edgecolor="white", width=0.62, error_kw=dict(lw=1.8))
        for i, b in enumerate(bars):
            ax.text(b.get_x()+b.get_width()/2, b.get_height()+errs[i]+0.35,
                    f"+{vals[i]:.1f}", ha="center", va="bottom", fontsize=14, fontweight="bold")
        ax.axhline(0, color="#666", lw=1)
        ax.set_xticks(range(len(arms))); ax.set_xticklabels([pretty[a] for a in arms])
        ax.set_ylabel(f"6-digit {metric} gap over baseline (pts)")
        ax.set_title(f"6-digit {metric} transfer gap")
        ax.set_ylim(0, max(v+e for v, e in zip(vals, errs)) * 1.25)
    fig.suptitle("Transfer gap spectrum at the decisive length  (mean ± seed std, 50–300 window)",
                 fontsize=16, fontweight="bold", y=1.0)
    fig.tight_layout()
    fig.savefig(os.path.join(OUTDIR, "tr_fig_gap_spectrum.png"), dpi=DPI, bbox_inches="tight")
    plt.close(fig)


def main():
    data = load_all_results()
    fig_trajectories(data, "EM")
    fig_trajectories(data, "PD")
    fig_6dig_focus(data)
    fig_indist_mastery(data)
    #fig_loss(load_loss())
    fig_gap_spectrum(load_gaps())
    print("done.")


if __name__ == "__main__":
    main()