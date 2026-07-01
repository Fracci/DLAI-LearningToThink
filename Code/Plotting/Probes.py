"""
Probes.py — generates publication-quality figures from
probe_results.xlsx (the aggregated output of Rule30Probe/RolloutProbe/
CarryOnlyProbe's layer-sweep CSVs, compiled into one workbook). No probing or
model logic runs here — this is a pure read-xlsx / render-matplotlib script.

Reads the unified probe layer-sweep results and produces four figures:
  1. gap_by_layer        : the headline curve — gap-over-floor vs depth, all probes.
  2. trained_vs_floor    : per-probe trained accuracy vs random floor (small multiples).
  3. peak_gap_bars       : peak gap per probe, labelled with the layer it peaks at.
  4. carry_focus         : carry_in & gen_dist trained/floor/gap (the long-range story).
"""
import os
from collections import OrderedDict
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
from openpyxl import load_workbook

XLSX = "probe_results.xlsx"
INDIR = "Results"
OUTDIR = "Plots/Probes"
DPI = 300

# consistent styling 
plt.rcParams.update({
    "font.family": "DejaVu Sans",
    "font.size": 12,
    "axes.titlesize": 14,
    "axes.titleweight": "bold",
    "axes.labelsize": 12,
    "axes.spines.top": False,
    "axes.spines.right": False,
    "axes.grid": True,
    "grid.alpha": 0.25,
    "grid.linewidth": 0.8,
    "legend.frameon": False,
    "figure.dpi": 120,
})

# label -> (color, linestyle, marker). Local feature = warm/dashed; long-range = cool/solid.
STYLE = OrderedDict([
    ("Rule30 · neighborhood (local)",        ("#d1495b", "--", "o")),
    ("Carry · carry_in (state)",             ("#2e7d32", "-",  "s")),
    ("Carry · gen_dist (distance)",          ("#1b9e77", "-",  "D")),
    ("Rollout · cell_above",                 ("#3060c0", "-",  "^")),
    ("Rollout · neighborhood",               ("#6a3d9a", "-",  "v")),
])

# map (model, target) from the sheet -> pretty label above
KEY = {
    ("Rule30", "neighborhood"):  "Rule30 · neighborhood (local)",
    ("Carry", "carry_in"):       "Carry · carry_in (state)",
    ("Carry", "gen_dist"):       "Carry · gen_dist (distance)",
    ("Rollout", "cell_above"):   "Rollout · cell_above",
    ("Rollout", "neighborhood"): "Rollout · neighborhood",
}


def load_data():
    """Read the All_Probes sheet of probe_results.xlsx into {pretty_label: {layer, trained, floor, gap, chance}},
    mapping only the (model, target) pairs listed in KEY."""

    wb = load_workbook(os.path.join(INDIR, XLSX))
    ws = wb["All_Probes"]
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}
    data = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        model = row[idx["Model"]]; target = row[idx["Probe target"]]
        label = KEY.get((model, target))
        if label is None:
            continue

        d = data.setdefault(label, {"layer": [], "trained": [], "floor": [], "gap": [], "chance": []})
        d["layer"].append(int(row[idx["Layer"]]))
        d["trained"].append(float(row[idx["Trained %"]]))
        d["floor"].append(float(row[idx["Random floor %"]]))
        d["gap"].append(float(row[idx["Gap (pts)"]]))
        d["chance"].append(float(row[idx["Chance %"]]))

    # sort each by layer
    for d in data.values():
        order = sorted(range(len(d["layer"])), key=lambda i: d["layer"][i])
        for k in d:
            d[k] = [d[k][i] for i in order]
    return data


def fig_gap_by_layer(data):
    """Headline figure: gap-over-floor vs transformer layer for every probe in
    STYLE, with a star marking each curve's peak layer — the local-early vs.
    long-range-deep dissociation in one plot."""

    fig, ax = plt.subplots(figsize=(9, 5.6))
    ax.axhline(0, color="#666", lw=1, zorder=1)

    for label, (color, ls, mk) in STYLE.items():
        if label not in data:
            continue

        d = data[label]
        ax.plot(d["layer"], d["gap"], ls=ls, marker=mk, color=color,
                lw=2.4, ms=7, label=label, zorder=3)
        
        # mark the peak
        pk = max(range(len(d["gap"])), key=lambda i: d["gap"][i])
        ax.scatter([d["layer"][pk]], [d["gap"][pk]], s=240, marker="*",
                   color=color, edgecolor="white", linewidth=1.2, zorder=4)
        
    ax.set_xlabel("Transformer layer (depth →)")
    ax.set_ylabel("Probe gap over random floor  (pts)")
    ax.set_title("World-model decodability by depth\nlocal features peak early; long-range features peak deep")
    ax.xaxis.set_major_locator(MaxNLocator(integer=True))

    ax.legend(loc="upper left", fontsize=10.5)
    ax.annotate("★ = peak layer", xy=(0.985, 0.03), xycoords="axes fraction",
                ha="right", va="bottom", fontsize=10, color="#444")
    
    fig.tight_layout()
    fig.savefig(os.path.join(OUTDIR, "fig_gap_by_layer.png"), dpi=DPI, bbox_inches="tight")
    plt.close(fig)


def fig_trained_vs_floor(data):
    """Small-multiples grid: one panel per probe, trained accuracy vs random
    floor by layer, with the gap shaded — makes the "gap, not raw accuracy"
    framing visually explicit for each probe individually."""

    labels = [l for l in STYLE if l in data]
    n = len(labels)
    cols = 3
    rows = (n + cols - 1) // cols
    fig, axes = plt.subplots(rows, cols, figsize=(13, 4.2 * rows), squeeze=False)

    for ax in axes.flat:
        ax.set_visible(False)

    for i, label in enumerate(labels):
        ax = axes[i // cols][i % cols]; ax.set_visible(True)
        color = STYLE[label][0]; d = data[label]
        ax.plot(d["layer"], d["trained"], "-o", color=color, lw=2.4, ms=6, label="trained")
        ax.plot(d["layer"], d["floor"], "--s", color="#888", lw=2, ms=5, label="random floor")
        ax.fill_between(d["layer"], d["floor"], d["trained"],
                        where=[t >= f for t, f in zip(d["trained"], d["floor"])],
                        color=color, alpha=0.12)
        
        if d["chance"]:
            ax.axhline(d["chance"][0], color="#bbb", ls=":", lw=1.2, label="chance")
        ax.set_title(label, fontsize=12)
        ax.set_xlabel("layer"); ax.set_ylabel("accuracy (%)")
        ax.set_ylim(0, 103)
        ax.xaxis.set_major_locator(MaxNLocator(integer=True))
        ax.legend(fontsize=9, loc="best")

    fig.suptitle("Trained accuracy vs random floor — the shaded gap is the result",
                 fontsize=15, fontweight="bold", y=1.0)
    fig.tight_layout()
    fig.savefig(os.path.join(OUTDIR, "fig_trained_vs_floor.png"), dpi=DPI, bbox_inches="tight")
    plt.close(fig)


def fig_peak_gap_bars(data):
    """Bar chart of each probe's single best (peak) gap over its floor,
    annotated with the layer where that peak occurs — a compact ranking view
    that also sets up the "strongest probe ≠ best transfer" dissociation."""

    labels = [l for l in STYLE if l in data]
    peaks = []; peak_layers = []; colors = []

    for label in labels:
        d = data[label]
        pk = max(range(len(d["gap"])), key=lambda i: d["gap"][i])
        peaks.append(d["gap"][pk]); peak_layers.append(d["layer"][pk]); colors.append(STYLE[label][0])

    fig, ax = plt.subplots(figsize=(9, 5.2))
    bars = ax.bar(range(len(labels)), peaks, color=colors, edgecolor="white", linewidth=1.2)

    for i, (b, L) in enumerate(zip(bars, peak_layers)):
        ax.text(b.get_x() + b.get_width() / 2, b.get_height() + 0.8,
                f"{peaks[i]:.1f}\n(L{L})", ha="center", va="bottom", fontsize=10.5, fontweight="bold")
        
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels([l.replace(" · ", "\n") for l in labels], fontsize=10)
    ax.set_ylabel("Peak gap over floor (pts)")
    ax.set_title("Peak world-model decodability per probe\n(strongest probe ≠ best transfer — see report)")
    ax.set_ylim(0, max(peaks) * 1.18)

    fig.tight_layout()
    fig.savefig(os.path.join(OUTDIR, "fig_peak_gap_bars.png"), dpi=DPI, bbox_inches="tight")
    plt.close(fig)


def fig_carry_focus(data):
    """Carry-arm-specific figure: trained/floor curves (left) and gap curves
    (right) for carry_in and gen_dist side by side — the long-range,
    depth-assembled feature story in isolation. No-op if neither key is present."""

    keys = ["Carry · carry_in (state)", "Carry · gen_dist (distance)"]
    keys = [k for k in keys if k in data]

    if not keys:
        return
    fig, (axL, axR) = plt.subplots(1, 2, figsize=(13, 5))

    for label in keys:
        c = STYLE[label][0]; mk = STYLE[label][2]; d = data[label]
        axL.plot(d["layer"], d["trained"], "-", marker=mk, color=c, lw=2.4, ms=6, label=f"{label.split('·')[1].strip()} — trained")
        axL.plot(d["layer"], d["floor"], "--", marker=mk, color=c, lw=1.6, ms=4, alpha=0.55, label=f"{label.split('·')[1].strip()} — floor")
        axR.plot(d["layer"], d["gap"], "-", marker=mk, color=c, lw=2.6, ms=7, label=label.split("·")[1].strip())

    axL.set_title("Carry latents: trained vs floor"); axL.set_xlabel("layer"); axL.set_ylabel("accuracy (%)")
    axL.xaxis.set_major_locator(MaxNLocator(integer=True)); axL.legend(fontsize=9)

    axR.axhline(0, color="#666", lw=1)
    axR.set_title("Carry latents: gap over floor\n(near-zero early = long-range feature not yet computed)")
    axR.set_xlabel("layer"); axR.set_ylabel("gap (pts)")
    axR.xaxis.set_major_locator(MaxNLocator(integer=True)); axR.legend(fontsize=10)
    
    fig.tight_layout()
    fig.savefig(os.path.join(OUTDIR, "fig_carry_focus.png"), dpi=DPI, bbox_inches="tight")
    plt.close(fig)


def main():
    """Load probe_results.xlsx and generate all four figures. Pure plotting
    entry point — no probing runs from this file."""
    
    data = load_data()
    fig_gap_by_layer(data)
    fig_trained_vs_floor(data)
    fig_peak_gap_bars(data)
    fig_carry_focus(data)
    print("saved: fig_gap_by_layer.png, fig_trained_vs_floor.png, "
          "fig_peak_gap_bars.png, fig_carry_focus.png")


if __name__ == "__main__":
    main()